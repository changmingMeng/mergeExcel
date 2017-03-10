# encoding: utf-8

import xlrd
import csv
from openpyxl import Workbook
from openpyxl import load_workbook
import chardet

class MakeData3G(object):


    def read_file(self, filename):
        self.readfilename = filename

    def write_file(self, filename):
        namedict = {0: "时间",
                    1: "基站名称",
                    2: "小区名",
                    3: "CI",
                    4: "LAC",
                    5: "RNC名称",
                    6: "主下行扰码",
                    7: "CS域话务量  (含切)[erl]",
                    8: "数据流量(KByte)[kbps]",
                   }

        with xlrd.open_workbook(self.readfilename) as workbook:
            filesheet = workbook.sheet_by_index(0)
            filerown = filesheet.nrows
            filecoln = filesheet.ncols

            wb = Workbook()
            ws = wb.active
            if not filecoln == 13:
                raise ("file type error")
            for r in range(filerown):
                if r == 0:
                    for c in range(len(namedict.keys())):
                        ws.cell(column=c + 1, row=r + 1, value=namedict[c])
                else:
                    for c in range(len(namedict.keys())):
                        if c == 1:#B列是原表D
                            ws.cell(column=c + 1, row=r + 1, value=filesheet.cell_value(r, 3))
                        elif c == 2:#C列是原表E列
                            ws.cell(column=c + 1, row=r + 1, value=filesheet.cell_value(r, 4))
                        elif c == 3:#D列是原表C列
                            ws.cell(column=c + 1, row=r + 1, value=filesheet.cell_value(r, 2))
                        elif c == 4:#E列是原表F列
                            ws.cell(column=c + 1, row=r + 1, value=filesheet.cell_value(r, 5))
                        elif c == 5:#F列是原表B列
                            ws.cell(column=c + 1, row=r + 1, value=filesheet.cell_value(r, 1))
                        elif c == 6:#G列是空列
                            ws.cell(column=c + 1, row=r + 1, value="")
                        elif c == 7:#H列是总话务量=原表(H+I)*24
                            ws.cell(column=c + 1, row=r + 1, value=(filesheet.cell_value(r, 7)+
                                                                    filesheet.cell_value(r, 8))*24)
                        elif c == 8:#I列是总流量=原表J+K+(L+M)*1024
                            ws.cell(column=c + 1, row=r + 1, value=filesheet.cell_value(r, 9)+
                                                                    filesheet.cell_value(r, 10)+
                                                                   (filesheet.cell_value(r, 11)+
                                                                    filesheet.cell_value(r, 12))*1024)
                        else:
                            ws.cell(column=c + 1, row=r + 1, value=filesheet.cell_value(r, c))
        wb.save(filename)

def makedata_3G(inputfile, outputfile):
    ad = MakeData3G()
    ad.read_file(inputfile)
    ad.write_file(outputfile)

def test1():
    makedata_3G(r"W网监控常用指标X-20170307060016.xlsx".decode("utf-8").encode("GBK"),
              r"Wtest2.xlsx")


if __name__ == "__main__":
    test1()