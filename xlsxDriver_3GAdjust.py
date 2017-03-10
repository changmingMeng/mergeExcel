# encoding: utf-8

import xlrd
import csv
from openpyxl import Workbook
from openpyxl import load_workbook
import chardet

class Adjust3G(object):


    def read_file(self, filename):
        self.readfilename = filename

    def read_to_dict(self, filename):
        self.facdict = {}
        csvFile = file(filename, 'rb')
        csvReader = csv.reader(csvFile)

        # 第一行是表头不要,并且检测文件编码是否为utf-8

        #csvReader.next()[13].decode("GBK")
        # if not chardet.detect(csvReader.next()[13])["encoding"] == "utf-8":
        #     raise ("file encoding error:"+chardet.detect(csvReader.next()[13])["encoding"])

        for row in csvReader:
            self.facdict[row[1]] = row[0]

    def write_file(self, filename):
        namedict = {0: "时间",
                    1: "LAC",
                    2: "CI",
                    3: "小区名",
                    4: "NodeB名称",
                    5: "RNC名称",
                    6: "CS域话务量  (含切)[erl]",
                    7: "VS.RB.AMRWB.DL.12.65",
                    8: "PSR99业务上行忙时数据吞吐量[KByte]",
                    9: "PSR99业务下行忙时数据吞吐量[KByte]",
                    10: "HSUPA RLC层上行流量(KByte)[MByte]",
                    11: "HSDPA RLC层下行流量(KByte)[KByte]",
                    12: "无线接通率[%]",
                    13: "CS域AMR掉话率[%]",
                    14: "RAB建立拥塞率[%]",
                    15: "RRC建立成功率[%]",
                    16: "厂家"}

        with xlrd.open_workbook(self.readfilename) as workbook:
            filesheet = workbook.sheet_by_index(0)
            filerown = filesheet.nrows
            filecoln = filesheet.ncols

            print filesheet.cell_value(1, 4), filesheet.cell_value(1, 4).encode("utf-8"), \
                type(filesheet.cell_value(1, 4)), type(filesheet.cell_value(1, 4).encode("utf-8"))
            #print chardet.detect(filesheet.cell_value(1, 4))
            wb = Workbook()
            ws = wb.active
            if not filecoln == 13:
                raise ("file type error")
            for r in range(filerown):
                if r == 0:
                    for c in range(filecoln + 4):
                        ws.cell(column=c + 1, row=r + 1, value=namedict[c])
                else:
                    for c in range(filecoln+4):
                        if c == 1:#B列和F列交换
                            ws.cell(column=c + 1, row=r + 1, value=filesheet.cell_value(r, 5))
                        elif c == 5:#B列和F列交换
                            ws.cell(column=c + 1, row=r + 1, value=filesheet.cell_value(r, 1))
                        elif c in [6, 7, 8, 9 ,10 ,11]:#删除G列
                            if c in [6, 7]:#H,I列数值为小时平均值，要乘以24才是日话务量
                                ws.cell(column=c + 1, row=r + 1, value=filesheet.cell_value(r, c + 1) * 24)
                            elif c in [10, 11]:#L,M列单位为M，要转换成K
                                ws.cell(column=c + 1, row=r + 1, value=filesheet.cell_value(r, c + 1) * 1024)
                            else:
                                ws.cell(column=c + 1, row=r + 1, value=filesheet.cell_value(r, c+1))
                        elif c in [12, 13, 14, 15]:#新增一列“厂家”
                            ws.cell(column=c + 1, row=r + 1, value="")
                        elif c == 16:
                            key = filesheet.cell_value(r, 3).encode("utf-8")
                            #print key,self.facdict[key], type(key), chardet.detect(key)
                            try:
                                ws.cell(column=c + 1, row=r + 1, value=self.facdict[key])
                            except KeyError:
                                ws.cell(column=c + 1, row=r + 1, value="N/A")
                        else:
                            ws.cell(column=c + 1, row=r + 1, value=filesheet.cell_value(r, c))
        wb.save(filename)

def adjust_3G(inputfile, dictfile, outputfile):
    ad = Adjust3G()
    ad.read_file(inputfile)
    ad.read_to_dict(dictfile)
    ad.write_file(outputfile)

def test1():
    adjust_3G(r"W网监控常用指标X-20170307060016.xlsx".decode("utf-8").encode("GBK"),
              r"Query1.csv",
              r"Wtest1.xlsx")


if __name__ == "__main__":
    test1()