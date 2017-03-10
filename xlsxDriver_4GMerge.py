# encoding: utf-8

import xlrd
from openpyxl import Workbook
from openpyxl import load_workbook
from functools import wraps

import utils

class ExcelDriver(object):


    def read_file_1(self,filename1):
        self.readfilename1 = filename1

    def read_file_2(self, filename2):
        self.readfilename2 = filename2

    def write_file(self, writefile):
        self.writefile = writefile

    def write(self):
        ut = utils.Utils()
        with xlrd.open_workbook(self.readfilename2) as workbook2:
            file2sheet = workbook2.sheet_by_index(0)
            file2rown = file2sheet.nrows
            file2coln = file2sheet.ncols

        with xlrd.open_workbook(self.readfilename1) as workbook1:
            file1sheet = workbook1.sheet_by_index(0)
            file1rown = file1sheet.nrows
            file1coln = file1sheet.ncols

            print "file1 rn=", file1rown, "cn=", file1coln
            print "file2 rn=", file2rown, "cn=", file2coln
            #file1coln=30
            #file2coln=30
            wb = Workbook()
            ws1 = wb.active
            if file2coln == 30:
                for r in range(file2rown):
                    for c in range(file2coln):
                        if r > 0 and c == 0:
                            date1 = ut.tuple2Sqlite3Timestring(xlrd.xldate_as_tuple(file2sheet.cell_value(r, c), 0)[:3])
                            ws1.cell(column=c + 1, row=r + 1, value=date1)
                        else:
                            ws1.cell(column=c+1, row=r+1, value=file2sheet.cell_value(r, c))
            elif file2coln == 31:
                for r in range(file2rown):
                    for c in range(file2coln-1):
                        if r > 0 and c > 26:
                            ws1.cell(column=c+1, row=r+1, value=file2sheet.cell_value(r, c+1))
                        elif r > 0 and c == 0:
                            date1 = ut.tuple2Sqlite3Timestring(xlrd.xldate_as_tuple(file2sheet.cell_value(r, c), 0)[:3])
                            print "date1=", date1
                            ws1.cell(column=c+1, row=r+1, value=date1)
                        else:
                            ws1.cell(column=c+1, row=r+1, value=file2sheet.cell_value(r, c))
            else:
                raise ("excel format error")

            for r in range(file1rown):
                if r is not 0:
                    for c in range(file1coln):
                        # if c == 0:
                        #     date2 = tuple([int(x) for x in file1sheet.cell_value(r, c).split('-')])
                        #     print "date2=", date2
                        #     ws1.cell(column=c + 1, row=r + file2rown, value=date2)
                        # else:
                            ws1.cell(column=c+1, row=r+file2rown, value=file1sheet.cell_value(r, c))

            wb.save(self.writefile)




def copy_excel(f1, f2, f3):
    ed = ExcelDriver()
    ed.read_file_1(f1)
    ed.read_file_2(f2)
    ed.write_file(f3)
    ed.write()

def test1():
    copy_excel(
        r"E:\developtools\PyCharm 2016.3\projects\mergeExcel\LTE小区级日数据流量及忙时KPI指标_20170306.xlsx".decode("utf-8").encode(
            "GBK"),
        r"E:\developtools\PyCharm 2016.3\projects\mergeExcel\LTE小区级日数据流量及忙时KPI指标备份_20170306.xlsx".decode(
            "utf-8").encode("GBK"),
        r"E:\developtools\PyCharm 2016.3\projects\mergeExcel\LTE小区级日数据流量及忙时KPI指标备份_20170306.xlsx".decode(
            "utf-8").encode("GBK"))

def test2():
    copy_excel(
        r"E:\developtools\PyCharm 2016.3\projects\mergeExcel\LTE小区级日数据流量及忙时KPI指标_20170304.xlsx".decode("utf-8").encode(
            "GBK"),
        r"E:\developtools\PyCharm 2016.3\projects\mergeExcel\LTE小区级日数据流量及忙时KPI指标备份_20170304.xlsx".decode(
            "utf-8").encode("GBK"),
        r"E:\developtools\PyCharm 2016.3\projects\mergeExcel\test4.xlsx")

def test3():
    copy_excel(r"E:\developtools\PyCharm 2016.3\projects\mergeExcel\test1.xlsx",
                r"E:\developtools\PyCharm 2016.3\projects\mergeExcel\test2.xlsx",
               r"E:\developtools\PyCharm 2016.3\projects\mergeExcel\test2.xlsx")

def log_time_delta(func):
    @wraps(func)
    def deco():
        start = datetime.now()
        res = func()
        end = datetime.now()
        delta = end - start
        print("func runed ", delta)
        return res
    return deco

if __name__ == "__main__":
    test1()
